#!/usr/bin/env python3
"""
debug_tag_audit.py — find issue-tagged debug prints that outlived their issue.

Every fix logged in Spreadsheets/Issue_Log.xlsx adds temporary console prints
tagged `ISSUE #<n>` so the user can confirm from the Godot console that the
running build contains the fix. Once the user marks that row `Issue Resolved`,
those prints are dead noise and must be removed.

The codebase IS the registry: one ripgrep-equivalent pass over the ~60 .gd files
gives the exact set of issue numbers that still have live prints, so the cost of
this audit is flat no matter how many thousands of rows the log grows to. There
is no hand-maintained list to drift out of sync.

Usage:
    py -3 Scripts/Utilities/debug_tag_audit.py           # report only
    py -3 Scripts/Utilities/debug_tag_audit.py --write   # + stamp column J of Game_Issues

Report is printed and also written to Spreadsheets/_debug_tag_audit.txt.
Exit code 1 if any cleanup is required, else 0.
"""

import os
import re
import sys
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
WORKBOOK = os.path.join(ROOT, "Spreadsheets", "Issue_Log.xlsx")
REPORT = os.path.join(ROOT, "Spreadsheets", "_debug_tag_audit.txt")

SHEET = "Game_Issues"
COL_NUM = 1     # A  #
COL_STATUS = 7  # G  Status
COL_DEBUG = 10  # J  Debug (written by this script — never hand-edit)

SKIP_DIRS = {".godot", ".git", "addons", "Spreadsheets"}

TAG_RE = re.compile(r"ISSUE\s*#\s*(\d+)", re.IGNORECASE)
PRINT_RE = re.compile(r"\b(print|print_rich|printerr|print_debug|push_error|push_warning)\s*\(")
COMMENT_RE = re.compile(r"^\s*#")

# Statuses whose prints are still doing their job — leave them alone.
KEEP_STATUSES = {"New Issue", "Fix Failed", "Improve Further", "New Functionality",
                 "Untested Fix", "Manual Fix"}
DONE_STATUSES = {"Issue Resolved"}


def scan_code():
    """-> {issue_num: {"prints": [hit...], "comments": [hit...]}}, hit = dict."""
    found = defaultdict(lambda: {"prints": [], "comments": []})
    for dirpath, dirnames, filenames in os.walk(ROOT):
        dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS]
        for fn in filenames:
            if not fn.endswith(".gd"):
                continue
            path = os.path.join(dirpath, fn)
            rel = os.path.relpath(path, ROOT).replace("\\", "/")
            try:
                lines = open(path, encoding="utf-8").read().splitlines()
            except (OSError, UnicodeDecodeError):
                continue
            for i, line in enumerate(lines, start=1):
                m = TAG_RE.search(line)
                if not m:
                    continue
                num = int(m.group(1))
                is_comment = bool(COMMENT_RE.match(line))
                is_print = bool(PRINT_RE.search(line)) and not is_comment
                hit = {
                    "file": rel,
                    "line": i,
                    "text": line.strip(),
                    # Deleting this line alone would empty an if/for/else block ->
                    # parse error. Needs the surrounding block removed or a `pass`.
                    "block_risk": _block_risk(lines, i),
                }
                found[num]["prints" if is_print else "comments"].append(hit)
    return found


def _block_risk(lines, lineno):
    """True if this print looks like the only statement in its block."""
    idx = lineno - 1
    cur = lines[idx]
    indent = len(cur) - len(cur.lstrip())
    # previous non-blank line opens a block at a shallower indent?
    j = idx - 1
    while j >= 0 and not lines[j].strip():
        j -= 1
    if j < 0:
        return False
    prev = lines[j]
    prev_indent = len(prev) - len(prev.lstrip())
    if not (prev.rstrip().endswith(":") and prev_indent < indent):
        return False
    # next non-blank line — if it dedents, this print was the whole block
    k = idx + 1
    while k < len(lines) and not lines[k].strip():
        k += 1
    if k >= len(lines):
        return True
    nxt = lines[k]
    return (len(nxt) - len(nxt.lstrip())) < indent


def read_log():
    """-> ({issue_num: status}, worksheet, workbook)"""
    import openpyxl
    wb = openpyxl.load_workbook(WORKBOOK)
    ws = wb[SHEET]
    statuses = {}
    for r in range(2, ws.max_row + 1):
        num = ws.cell(r, COL_NUM).value
        if num is None:
            continue
        try:
            num = int(str(num).strip().lstrip("#"))
        except ValueError:
            continue
        statuses[num] = (str(ws.cell(r, COL_STATUS).value or "").strip(), r)
    return statuses, ws, wb


def main():
    write = "--write" in sys.argv
    code = scan_code()
    statuses, ws, wb = read_log()

    cleanup, orphans, keeping = [], [], []
    for num in sorted(code):
        prints = code[num]["prints"]
        if not prints:
            continue
        entry = statuses.get(num)
        if entry is None:
            orphans.append((num, prints))
        elif entry[0] in DONE_STATUSES:
            cleanup.append((num, entry[0], prints))
        else:
            keeping.append((num, entry[0], len(prints)))

    out = []
    w = out.append
    w("=" * 78)
    w("ISSUE DEBUG-TAG AUDIT")
    w("=" * 78)
    w(f"issues with tags in code : {len(code)}")
    w(f"issues with live prints  : {sum(1 for n in code if code[n]['prints'])}")
    w(f"rows in {SHEET}          : {len(statuses)}")
    w("")

    w("-" * 78)
    w(f"[1] CLEANUP REQUIRED — resolved issues still printing ({len(cleanup)} issues, "
      f"{sum(len(p) for _, _, p in cleanup)} lines)")
    w("-" * 78)
    if not cleanup:
        w("  none — every resolved issue is clean.")
    for num, status, prints in cleanup:
        w(f"\n  ISSUE #{num}  [{status}]  {len(prints)} print line(s)")
        for h in prints:
            flag = "  <-- BLOCK RISK: sole statement in its block, needs care" if h["block_risk"] else ""
            w(f"    {h['file']}:{h['line']}{flag}")
            w(f"        {h['text'][:150]}")
    w("")

    w("-" * 78)
    w(f"[2] ORPHAN TAGS — tag number not present in the log ({len(orphans)})")
    w("-" * 78)
    if not orphans:
        w("  none.")
    for num, prints in orphans:
        w(f"  ISSUE #{num}: {len(prints)} print(s) — " +
          ", ".join(f"{h['file']}:{h['line']}" for h in prints[:6]))
    w("")

    w("-" * 78)
    w(f"[3] KEEP — prints on issues still open/untested ({len(keeping)})")
    w("-" * 78)
    w("  " + ", ".join(f"#{n} ({s}, {c})" for n, s, c in keeping) if keeping else "  none.")
    w("")
    w("NOTE: `# ISSUE #n ...` comments are permanent documentation and are NOT")
    w("      reported here. Only executable print/push_error lines are.")

    text = "\n".join(out)
    print(text)
    with open(REPORT, "w", encoding="utf-8") as f:
        f.write(text + "\n")
    print(f"\n[report written to {os.path.relpath(REPORT, ROOT)}]")

    if write:
        if ws.cell(1, COL_DEBUG).value != "Debug":
            ws.cell(1, COL_DEBUG).value = "Debug"
        for num, (status, row) in statuses.items():
            n = len(code.get(num, {}).get("prints", []))
            if n and status in DONE_STATUSES:
                ws.cell(row, COL_DEBUG).value = f"Live: {n} (REMOVE)"
            elif n:
                ws.cell(row, COL_DEBUG).value = f"Live: {n}"
            else:
                ws.cell(row, COL_DEBUG).value = None
        wb.save(WORKBOOK)
        print(f"[column J of {SHEET} updated]")

    return 1 if cleanup else 0


if __name__ == "__main__":
    sys.exit(main())
